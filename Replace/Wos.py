from openpyxl import load_workbook
from Class import WOS_Library

def Wos(path):
    #открытие исходного файла
    wb = load_workbook(path)
    ws = wb.active
    
    #список статей SCOPUS
    all_wos_list_library = []

    #кол-во ФИО авторов
    count_all_author = 0
    
    #ПАРСИНГ КАЖДОЙ СТРОКИ
    for row in ws.iter_rows(min_row=2, values_only=True):
        pass