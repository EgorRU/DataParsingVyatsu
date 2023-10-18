from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl import Workbook
import json
from Logging import writeFile


def match(text, alphabet=set('абвгдеёжзийклмнопрстуфхцчшщъыьэюя')):
    return not alphabet.isdisjoint(text.lower())


def full_description(obj):
    s = f"{obj.original_author}, {obj.title} - {obj.year}. "
    if hasattr(obj, "volume") and obj.volume != None and obj.volume!="":
        s += f" - Vol. {obj.volume}. "
    if hasattr(obj, "issue") and obj.issue != None and obj.issue != "":
        s += f" - №{obj.issue}. "
    if hasattr(obj, "start_page") and obj.start_page != None and obj.start_page!="":
        s += f" - pp. {obj.start_page}- "
    if hasattr(obj, "end_page") and obj.end_page != None and obj.end_page!="":
        s += f"{obj.end_page}. "
    else:
        s += "  "
    if hasattr(obj, "number") and obj.number != None and obj.number != "":
        s += f" - Vol. {obj.number}. "
    if hasattr(obj, "source_name") and obj.source_name != None and obj.source_name != "":
        s += f"source_name: {obj.source_name} "
    if hasattr(obj, "title_article") and obj.title_article != None and obj.title_article != "":
        s += f"title: {obj.title_article} "
    if hasattr(obj, "pages") and obj.pages != None and obj.pages != "":
        s += f"pages: {obj.pages} "
    if hasattr(obj, "publisher") and obj.publisher != None and obj.publisher != "":
        s += f"publisher: {obj.publisher} "
    return s 


def Upload(path, list_new, list_ident, list_remove):
    # создание нового файла
    wb = Workbook()
    ws = wb.active
    
    #границы ячейки
    border = Border(top=Side(border_style='thick',color="FFFF00"))

    #название листа 
    ws.title = "Сравнение данных"
        
    #получаем заголовки столбиков в список
    list_members = []
    if len(list_new)>0:
        list_members = list_new[0].__dir__()
        if len(list_ident)>0:
            list_members = list(set(list_members).intersection(set(list_ident[0].__dir__())))
        if len(list_remove)>0:
            list_members = list(set(list_members).intersection(set(list_remove[0].__dir__()))) 
    else:
        if len(list_ident)>0:
            list_members = list_ident[0].__dir__()
            if len(list_remove)>0:
                list_members = list(set(list_members).intersection(set(list_remove[0].__dir__()))) 
        else:
            if len(list_remove)>0:
                list_members = list_remove[0].__dir__()
    
    #удаляем спец. поля класса
    new_list_members = []       
    for var in list_members:
        if var[0] != "_" and var[0].islower() and var[:5]!="clear":
            new_list_members.append(var)
            
    #формируем сортированный список
    list_members = []
    new_list_members.remove("author")
    new_list_members.remove("title")
    new_list_members.remove("year")
    new_list_members.remove("original_author")
    
    list_members = ["author", "original_author", "title", "year", "full_bibliographic_description"]
    
    if "full_bibliographic_description" in new_list_members:
        new_list_members.remove("full_bibliographic_description")
    
    #если есть поля, то добавляем в нужной очерёдности
    if "doi" in new_list_members:
        list_members.append("doi")
        new_list_members.remove("doi")
        
    if "link" in new_list_members:
        list_members.append("link")
        new_list_members.remove("link")
        
    if "volume" in new_list_members:
        list_members.append("volume")
        new_list_members.remove("volume")
        
    if "issue" in new_list_members:
        list_members.append("issue")
        new_list_members.remove("issue")
        
    if "start_page" in new_list_members:
        list_members.append("start_page")
        new_list_members.remove("start_page")
        
    if "end_page" in new_list_members:
        list_members.append("end_page")
        new_list_members.remove("end_page")
        
    if "number_of_pages" in new_list_members:
        list_members.append("number_of_pages")
        new_list_members.remove("number_of_pages")
        
    if "number_article" in new_list_members:
        list_members.append("number_article")
        new_list_members.remove("number_article")
        
    if "title_article" in new_list_members:
        list_members.append("title_article")
        new_list_members.remove("title_article")
        
    if "issn" in new_list_members:
        list_members.append("issn")
        new_list_members.remove("issn")

    if "eissn" in new_list_members:
        list_members.append("eissn")
        new_list_members.remove("eissn")
        
    #добавляем остальные поля
    new_list_members.sort()
    for i in range(len(new_list_members)):
        list_members.append(new_list_members[i])
        
    if "source" in new_list_members:
        list_members.remove("source")
        list_members.append("source")
        new_list_members.remove("source")

    #меняем заголовки столбиков
    ws.cell(row=1, column=1).value = "№№"
    ws.cell(row=1, column=2).value = "№"
    for index, val in enumerate(list_members):
        ws.cell(row=1, column=index+3).value = val
        ws.cell(row=1, column=index+3).font = Font(size=16)
    ws.cell(row=1, column=5).value = "title_article"
    for i in range(6, 20):
        if  ws.cell(row=1, column=i).value == "title_article":
            ws.cell(row=1, column=i).value = "title"
            break
    
    #меняем ширину ячеек
    for i in range(len(list_members)):
        ws.column_dimensions[ws.cell(row=1, column=i+3).column_letter].width = 60
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    if "original_author" in list_members:
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 90
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 200
        ws.column_dimensions["H"].width = 70    
        ws.column_dimensions["I"].width = 150
    else:
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 90
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 200
        ws.column_dimensions["G"].width = 70    
        ws.column_dimensions["H"].width = 150
      
    temp_row = 2
    count_article = 0
    
    #словари с кол-во статей для add, ident, remove листов
    dict_add_count = dict()
    dict_ident_count = dict()
    dict_remove_count = dict()
    
    #словари с кол-во статей для add, ident, remove листов
    dict_add_citated = dict()
    dict_ident_citated = dict()
    dict_remove_citated = dict()
    
    list_data = []
    try:
        #словарь с сотрудниками
        file = open('employee.json', 'r', encoding="utf-8")
        list_data = json.load(file)
    except:
        writeFile("exception", "Файл 'employee.json' не найден\n")
    
    count = 0
    #добавляем новые элементы
    title = ""
    #по новым элементам списка
    for i in range(len(list_new)):
        #берём текущее название
        title_temp = getattr(list_new[i], "title")
        #1-2 ячейки для нумерации
        ws.cell(row=i+temp_row, column=1).font = Font(size=16)
        ws.cell(row=i+temp_row, column=2).font = Font(size=16)
        ws.cell(row=i+temp_row, column=1).fill = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
        ws.cell(row=i+temp_row, column=2).fill = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
        #если текущее название не совпало с прошлым, то отделяем чертой ячейки
        if title_temp!= title:
            #заполяем словари
            #кол-во статей по годам
            if list_new[i].year in dict_add_count:
                dict_add_count[list_new[i].year] += 1
            else:
                dict_add_count[list_new[i].year] = 1  
            
            #если есть столбец с цитированием
            if "citation" in list_members:
                #кол-во цитирований по годам
                if list_new[i].year in dict_add_citated:
                    dict_add_citated[list_new[i].year] += int(list_new[i].citation)
                else:
                    dict_add_citated[list_new[i].year] = int(list_new[i].citation)
                
            #ставим линию в 1 и 2 колонке
            ws.cell(row=i+temp_row, column=1).border = border
            ws.cell(row=i+temp_row, column=2).border = border
            count_article += 1 #кол-во записей
            count += 1 #кол-во записей
            #Бежим по списку атрибутов класса
            for j in range(len(list_members)):
                employe = False
                if list_members[j]=="author":
                    #проходимся по всем сотрудникам вуза
                    for k in range(len(list_data)):
                        #если есть совпадения - то добавляем бордер и выходим из цикла
                        if list_new[i].author in list_data[k] or list_new[i].author == list_data[k]:
                            employe = True
                            break
                if list_members[j]=="full_bibliographic_description":
                    if not hasattr(list_new[i],'full_bibliographic_description'):
                        ws.cell(row=i+temp_row, column=j+3).value = full_description(list_new[i])
                    else:
                         ws.cell(row=i+temp_row, column=j+3).value = list_new[i].full_bibliographic_description
                #иначе обычное поле заполняем
                else:
                    ws.cell(row=i+temp_row, column=j+3).value = getattr(list_new[i], list_members[j])
                #делаем бордер и красим в зелёный цвет
                if employe:
                    ws.cell(row=i+temp_row, column=j+3).border = Border(top = Side(border_style='thick', color='8B0000'),
                                                                                right = Side(border_style='thick', color='8B0000'),
                                                                                bottom = Side(border_style='thick', color='8B0000'),
                                                                                left = Side(border_style='thick', color='8B0000'))
                else:
                    ws.cell(row=i+temp_row, column=j+3).border = border
                ws.cell(row=i+temp_row, column=j+3).fill = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
        #иначе если текущее название сходится с прошлым, то бордер делать не надо
        else:
            repetition = True
            if hasattr(list_new[i],'eid') and i>0 and list_new[i].eid!=list_new[i-1].eid:
                repetition = False
                #заполяем словари
                #кол-во статей по годам
                if list_new[i].year in dict_add_count:
                    dict_add_count[list_new[i].year] += 1
                else:
                    dict_add_count[list_new[i].year] = 1  
                #если есть столбец с цитированием
                if "citation" in list_members:
                    #кол-во цитирований по годам
                    if list_new[i].year in dict_add_citated:
                        dict_add_citated[list_new[i].year] += int(list_new[i].citation)
                    else:
                        dict_add_citated[list_new[i].year] = int(list_new[i].citation)
                #ставим линию в 1 и 2 колонке
                ws.cell(row=i+temp_row, column=1).border = border
                ws.cell(row=i+temp_row, column=2).border = border
                count_article += 1 #кол-во записей
                count += 1 #кол-во записей

            #по всем атрибутам класса
            for j in range(len(list_members)):
                if list_members[j]=="author":
                    #по всем сотрудникам вуза
                    for k in range(len(list_data)):
                        #если есть совпадения, то выделяем ячейку и выходим из цикла
                        if list_new[i].author in list_data[k] or list_new[i].author == list_data[k]:
                            ws.cell(row=i+temp_row, column=j+3).border = Border(top = Side(border_style='thick', color='8B0000'),right = Side(border_style='thick', color='8B0000'),bottom = Side(border_style='thick', color='8B0000'),left = Side(border_style='thick', color='8B0000'))
                            break
                if list_members[j]=="full_bibliographic_description":
                    if not hasattr(list_new[i],'full_bibliographic_description'):
                        ws.cell(row=i+temp_row, column=j+3).value = full_description(list_new[i])
                    else:
                         ws.cell(row=i+temp_row, column=j+3).value = list_new[i].full_bibliographic_description
                #иначе обычное поле заполняем
                else:
                    ws.cell(row=i+temp_row, column=j+3).value = getattr(list_new[i], list_members[j])
                #если статья другая, то ставим бордер
                if repetition==False:
                    ws.cell(row=i+temp_row, column=j+3).border = border
                #красим ячейку в зелёный
                ws.cell(row=i+temp_row, column=j+3).fill = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
        #нумеруем 1 и 2 колонку
        ws.cell(row=i+temp_row, column=1).value = count_article
        ws.cell(row=i+temp_row, column=2).value = count
        #меняем текущее название для дальнейшего бордера
        title = title_temp
    temp_row += len(list_new)


    count = 0
    #добавляем одинаковые элементы
    title = ""
    #по новым элементам списка
    for i in range(len(list_ident)):
        #берём текущее название
        title_temp = getattr(list_ident[i], "title")
        #1-2 ячейки для нумерации
        ws.cell(row=i+temp_row, column=1).font = Font(size=16)
        ws.cell(row=i+temp_row, column=2).font = Font(size=16)
        #если текущее название не совпало с прошлым, то отделяем чертой ячейки
        if title_temp!= title:
            #заполяем словари
            #кол-во статей по годам
            if list_ident[i].year in dict_ident_count:
                dict_ident_count[list_ident[i].year] += 1
            else:
                dict_ident_count[list_ident[i].year] = 1  
            
            #если есть столбец с цитированием
            if "citation" in list_members:
                #кол-во цитирований по годам
                if list_ident[i].year in dict_ident_citated:
                    dict_ident_citated[list_ident[i].year] += int(list_ident[i].citation)
                else:
                    dict_ident_citated[list_ident[i].year] = int(list_ident[i].citation)
                
            #ставим линию в 1 и 2 колонке
            ws.cell(row=i+temp_row, column=1).border = border
            ws.cell(row=i+temp_row, column=2).border = border
            count_article += 1 #кол-во записей
            count += 1 #кол-во записей
            #Бежим по списку атрибутов класса
            for j in range(len(list_members)):
                employe = False
                if list_members[j]=="author":
                    #проходимся по всем сотрудникам вуза
                    for k in range(len(list_data)):
                        #если есть совпадения - то добавляем бордер и выходим из цикла
                        if list_ident[i].author in list_data[k] or list_ident[i].author == list_data[k]:
                            ws.cell(row=i+temp_row, column=j+3).border = Border(top = Side(border_style='thick', color='8B0000'),right = Side(border_style='thick', color='8B0000'),bottom = Side(border_style='thick', color='8B0000'),left = Side(border_style='thick', color='8B0000'))
                            employe = True
                            break
                if list_members[j]=="full_bibliographic_description":
                    if not hasattr(list_ident[i],'full_bibliographic_description'):
                        ws.cell(row=i+temp_row, column=j+3).value = full_description(list_ident[i])
                    else:
                        ws.cell(row=i+temp_row, column=j+3).value = list_ident[i].full_bibliographic_description
                #иначе обычное поле заполняем
                else:
                    ws.cell(row=i+temp_row, column=j+3).value = getattr(list_ident[i], list_members[j])
                #делаем бордер
                if employe:
                   ws.cell(row=i+temp_row, column=j+3).border = Border(top = Side(border_style='thick', color='8B0000'),
                                                                                right = Side(border_style='thick', color='8B0000'),
                                                                                bottom = Side(border_style='thick', color='8B0000'),
                                                                                left = Side(border_style='thick', color='8B0000'))
                else:
                    ws.cell(row=i+temp_row, column=j+3).border = border
        #иначе если текущщее название свходится с прошлым, то бордер делать не надо
        else:
            repetition = True
            if hasattr(list_ident[i],'eid') and i>0 and list_ident[i].eid!=list_ident[i-1].eid:
                repetition = False
                #заполяем словари
                #кол-во статей по годам
                if list_ident[i].year in dict_add_count:
                    dict_add_count[list_ident[i].year] += 1
                else:
                    dict_add_count[list_ident[i].year] = 1  
                #если есть столбец с цитированием
                if "citation" in list_members:
                    #кол-во цитирований по годам
                    if list_ident[i].year in dict_add_citated:
                        dict_add_citated[list_ident[i].year] += int(list_ident[i].citation)
                    else:
                        dict_add_citated[list_ident[i].year] = int(list_ident[i].citation)
                #ставим линию в 1 и 2 колонке
                ws.cell(row=i+temp_row, column=1).border = border
                ws.cell(row=i+temp_row, column=2).border = border
                count_article += 1 #кол-во записей
                count += 1 #кол-во записей
            #по всем атрибутам класса
            for j in range(len(list_members)):
                if list_members[j]=="author":
                    #по всем сотрудникам вуза
                    for k in range(len(list_data)):
                        #если есть совпадения, то выделяем ячейку и выходим из цикла
                        if list_ident[i].author in list_data[k] or list_ident[i].author == list_data[k]:
                            ws.cell(row=i+temp_row, column=j+3).border = Border(top = Side(border_style='thick', color='8B0000'),right = Side(border_style='thick', color='8B0000'),bottom = Side(border_style='thick', color='8B0000'),left = Side(border_style='thick', color='8B0000'))
                            break
                if list_members[j]=="full_bibliographic_description":
                    if not hasattr(list_ident[i],'full_bibliographic_description'):
                        ws.cell(row=i+temp_row, column=j+3).value = full_description(list_ident[i])
                    else:
                        ws.cell(row=i+temp_row, column=j+3).value = list_ident[i].full_bibliographic_description
                #иначе обычное поле заполняем
                else:
                    ws.cell(row=i+temp_row, column=j+3).value = getattr(list_ident[i], list_members[j])
                #если статья другая, то ставим бордер
                if repetition==False:
                    ws.cell(row=i+temp_row, column=j+3).border = border
        #нумеруем 1 и 2 колонку
        ws.cell(row=i+temp_row, column=1).value = count_article
        ws.cell(row=i+temp_row, column=2).value = count
        #меняем текущее название для дальнейшего бордера
        title = title_temp
    temp_row += len(list_ident)


    count = 0
    #добавляем удалённые элементы
    title = ""
    #по новым элементам списка
    for i in range(len(list_remove)):
        #берём текущее название
        title_temp = getattr(list_remove[i], "title")
        #1-2 ячейки для нумерации
        ws.cell(row=i+temp_row, column=1).font = Font(size=16)
        ws.cell(row=i+temp_row, column=2).font = Font(size=16)
        ws.cell(row=i+temp_row, column=1).fill = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
        ws.cell(row=i+temp_row, column=2).fill = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
        #если текущее название не совпало с прошлым, то отделяем чертой ячейки
        if title_temp!= title:
            #заполяем словари
            #кол-во статей по годам
            if list_remove[i].year in dict_remove_count:
                dict_remove_count[list_remove[i].year] += 1
            else:
                dict_remove_count[list_remove[i].year] = 1  
            
            #если есть столбец с цитированием
            if "citation" in list_members:
                #кол-во цитирований по годам
                if list_remove[i].year in dict_remove_citated:
                    dict_remove_citated[list_remove[i].year] += int(list_remove[i].citation)
                else:
                    dict_remove_citated[list_remove[i].year] = int(list_remove[i].citation)
                    
            #ставим линию в 1 и 2 колонке
            ws.cell(row=i+temp_row, column=1).border = border
            ws.cell(row=i+temp_row, column=2).border = border
            count_article += 1 #кол-во записей
            count += 1 #кол-во записей
            #Бежим по списку атрибутов класса
            for j in range(len(list_members)):
                employe = False
                if list_members[j]=="author":
                    #проходимся по всем сотрудникам вуза
                    for k in range(len(list_data)):
                        #если есть совпадения - то добавляем бордер и выходим из цикла
                        if list_remove[i].author in list_data[k] or list_remove[i].author == list_data[k]:
                            employe = True
                            break
                if list_members[j]=="full_bibliographic_description":
                    if not hasattr(list_remove[i],'full_bibliographic_description'):
                        ws.cell(row=i+temp_row, column=j+3).value = full_description(list_remove[i])
                    else:
                        ws.cell(row=i+temp_row, column=j+3).value = list_remove[i].full_bibliographic_description
                #иначе обычное поле заполняем
                else:
                    ws.cell(row=i+temp_row, column=j+3).value = getattr(list_remove[i], list_members[j])
                #делаем бордер и красим в красный цвет
                if employe:
                    ws.cell(row=i+temp_row, column=j+3).border = Border(top = Side(border_style='thick', color='8B0000'),
                                                                                right = Side(border_style='thick', color='8B0000'),
                                                                                bottom = Side(border_style='thick', color='8B0000'),
                                                                                left = Side(border_style='thick', color='8B0000'))
                else:
                    ws.cell(row=i+temp_row, column=j+3).border = border
                ws.cell(row=i+temp_row, column=j+3).fill = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
        #иначе если текущщее название свходится с прошлым, то бордер делать не надо
        else:
            repetition = True
            if hasattr(list_remove[i],'eid') and i>0 and list_remove[i].eid!=list_remove[i-1].eid:
                repetition = False
                #заполяем словари
                #кол-во статей по годам
                if list_remove[i].year in dict_add_count:
                    dict_add_count[list_remove[i].year] += 1
                else:
                    dict_add_count[list_remove[i].year] = 1  
                #если есть столбец с цитированием
                if "citation" in list_members:
                    #кол-во цитирований по годам
                    if list_remove[i].year in dict_add_citated:
                        dict_add_citated[list_remove[i].year] += int(list_remove[i].citation)
                    else:
                        dict_add_citated[list_remove[i].year] = int(list_remove[i].citation)
                #ставим линию в 1 и 2 колонке
                ws.cell(row=i+temp_row, column=1).border = border
                ws.cell(row=i+temp_row, column=2).border = border
                count_article += 1 #кол-во записей
                count += 1 #кол-во записей
            #по всем атрибутам класса
            for j in range(len(list_members)):
                if list_members[j]=="author":
                    #по всем сотрудникам вуза
                    for k in range(len(list_data)):
                        #если есть совпадения, то выделяем ячейку и выходим из цикла
                        if list_remove[i].author in list_data[k] or list_remove[i].author == list_data[k]:
                            ws.cell(row=i+temp_row, column=j+3).border = Border(top = Side(border_style='thick', color='8B0000'),right = Side(border_style='thick', color='8B0000'),bottom = Side(border_style='thick', color='8B0000'),left = Side(border_style='thick', color='8B0000'))
                            break
                if list_members[j]=="full_bibliographic_description":
                    if not hasattr(list_remove[i],'full_bibliographic_description'):
                        ws.cell(row=i+temp_row, column=j+3).value = full_description(list_remove[i])
                    else:
                        ws.cell(row=i+temp_row, column=j+3).value = list_remove[i].full_bibliographic_description
                #иначе обычное поле заполняем
                else:
                    ws.cell(row=i+temp_row, column=j+3).value = getattr(list_remove[i], list_members[j])
                #если статья другая, то ставим бордер
                if repetition==False:
                    ws.cell(row=i+temp_row, column=j+3).border = border
                #красим ячейку в красный
                ws.cell(row=i+temp_row, column=j+3).fill = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
        #нумеруем 1 и 2 колонку
        ws.cell(row=i+temp_row, column=1).value = count_article
        ws.cell(row=i+temp_row, column=2).value = count
        #меняем текущее название для дальнейшего бордера
        title = title_temp
    temp_row += len(list_remove)


    #изменяем высоту ячейки
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 40


    #перенос строк
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
         for cell in row_cells:
            cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            cell.font = Font(size=16)
    
    #статистика в новый лист книги
    ws = wb.create_sheet('Статистика')
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30
    
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 30
    ws.column_dimensions["K"].width = 30
    ws.column_dimensions["L"].width = 30
    ws.column_dimensions["M"].width = 30
    ws.column_dimensions["N"].width = 30
    
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Кол-во статей по годам'
    
    ws.merge_cells('I1:N1')
    ws['I1'] = 'Кол-во цитирований по годам'
    
    ws.merge_cells('A2:B2')
    ws['A2'] = 'Добавленные статьи'
    
    ws.merge_cells('C2:D2')
    ws['C2'] = 'Одинаковые статьи'

    ws.merge_cells('E2:F2')
    ws['E2'] = 'Удалённые статьи'
    
    ws.merge_cells('I2:J2')
    ws['I2'] = 'Добавленные статьи'
    
    ws.merge_cells('K2:L2')
    ws['K2'] = 'Одинаковые статьи'

    ws.merge_cells('M2:N2')
    ws['M2'] = 'Удалённые статьи'
    
    ws['A3'] = 'Год'
    ws['C3'] = 'Год'
    ws['E3'] = 'Год'
    ws['I3'] = 'Год'
    ws['K3'] = 'Год'
    ws['M3'] = 'Год'
    
    ws['B3'] = 'Кол-во'
    ws['D3'] = 'Кол-во'
    ws['F3'] = 'Кол-во'
    ws['J3'] = 'Кол-во'
    ws['L3'] = 'Кол-во'
    ws['N3'] = 'Кол-во'
    
    i = 4
    for key in dict_add_count:
        ws.cell(row=i, column=1).value = key
        ws.cell(row=i, column=2).value = dict_add_count[key]
        i += 1
    
    i = 4
    for key in dict_ident_count:
        ws.cell(row=i, column=3).value = key
        ws.cell(row=i, column=4).value = dict_ident_count[key]
        i += 1

    i = 4
    for key in dict_remove_count:
        ws.cell(row=i, column=5).value = key
        ws.cell(row=i, column=6).value = dict_remove_count[key]
        i += 1
        
    i = 4
    for key in dict_add_citated:
        ws.cell(row=i, column=9).value = key
        ws.cell(row=i, column=10).value = dict_add_citated[key]
        i += 1
        
    i = 4
    for key in dict_ident_citated:
        ws.cell(row=i, column=11).value = key
        ws.cell(row=i, column=12).value = dict_ident_citated[key]
        i += 1
        
    i = 4
    for key in dict_remove_citated:
        ws.cell(row=i, column=13).value = key
        ws.cell(row=i, column=14).value = dict_remove_citated[key]
        i += 1
        

    #перенос строк и центирование текста
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
         for cell in row_cells:
            cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            cell.font = Font(size=16)
            
    wb.save(path)

    #выгружаем только сотрудников вуза
    wb = Workbook()
    ws = wb.active

    #границы ячейки
    border = Border(top=Side(border_style='thick',color="FFFF00"))

    #название листа 
    ws.title = "Сравнение данных"
    
    #меняем заголовки столбиков
    ws.cell(row=1, column=1).value = "№№"
    ws.cell(row=1, column=2).value = "№"
    for index, val in enumerate(list_members):
        ws.cell(row=1, column=index+3).value = val
        ws.cell(row=1, column=index+3).font = Font(size=16)
    ws.cell(row=1, column=5).value = "title_article"
    for i in range(6, 20):
        if  ws.cell(row=1, column=i).value == "title_article":
            ws.cell(row=1, column=i).value = "title"
            break

    #меняем ширину ячеек
    for i in range(len(list_members)):
        ws.column_dimensions[ws.cell(row=1, column=i+3).column_letter].width = 60
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    if "original_author" in list_members:
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 90
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 200
        ws.column_dimensions["H"].width = 70    
        ws.column_dimensions["I"].width = 150
    else:
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 90
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 200
        ws.column_dimensions["G"].width = 70    
        ws.column_dimensions["H"].width = 150
      
    temp_row = 2
    count_article = 0
    
    #словари с кол-во статей для add, ident, remove листов
    dict_add_count = dict()
    dict_ident_count = dict()
    dict_remove_count = dict()
    
    #словари с кол-во статей для add, ident, remove листов
    dict_add_citated = dict()
    dict_ident_citated = dict()
    dict_remove_citated = dict()

    count = 0
    #по новым элементам списка
    title = ""
    for i in range(len(list_new)):
        #сотрудник
        employe_flag = False
        employe = list_new[i].author
        for k in range(len(list_data)):
            if list_new[i].author in list_data[k] or list_new[i].author == list_data[k]:
                employe_flag = True
                break
        if employe_flag:
            #берём текущее название
            title_temp = getattr(list_new[i], "title")
            #1 и 2 ячейки для нумерации
            ws.cell(row=temp_row, column=1).font = Font(size=16)
            ws.cell(row=temp_row, column=2).font = Font(size=16)
            ws.cell(row=temp_row, column=1).fill = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
            ws.cell(row=temp_row, column=2).fill = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
            #если текущее название не совпало с прошлым, то отделяем чертой ячейки
            if title_temp!= title:
                #заполяем словари
                #кол-во статей по годам
                if list_new[i].year in dict_add_count:
                    dict_add_count[list_new[i].year] += 1
                else:
                    dict_add_count[list_new[i].year] = 1  
            
                #если есть столбец с цитированием
                if "citation" in list_members:
                    #кол-во цитирований по годам
                    if list_new[i].year in dict_add_citated:
                        dict_add_citated[list_new[i].year] += int(list_new[i].citation)
                    else:
                        dict_add_citated[list_new[i].year] = int(list_new[i].citation)
                
                #ставим линию в 1 и 2 колонке
                ws.cell(row=temp_row, column=1).border = border
                ws.cell(row=temp_row, column=2).border = border
                count_article += 1 #кол-во записей
                count += 1 #кол-во записей
                #Бежим по списку атрибутов класса
                for j in range(len(list_members)):
                    if list_members[j]=="full_bibliographic_description":
                        if not hasattr(list_new[i],'full_bibliographic_description'):
                            ws.cell(row=temp_row, column=j+3).value = full_description(list_new[i])
                        else:
                            ws.cell(row=temp_row, column=j+3).value = list_new[i].full_bibliographic_description
                    #иначе обычное поле заполняем
                    else:
                        ws.cell(row=temp_row, column=j+3).value = getattr(list_new[i], list_members[j])
                    #делаем бордер и красим в зелёный цвет
                    ws.cell(row=temp_row, column=j+3).border = border
                    ws.cell(row=temp_row, column=j+3).fill = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
            #иначе если текущее название сходится с прошлым, то бордер делать не надо
            else:
                repetition = True
                if hasattr(list_new[i],'eid') and i>0 and list_new[i].eid!=list_new[i-1].eid:
                    repetition = False
                    #заполяем словари
                    #кол-во статей по годам
                    if list_new[i].year in dict_add_count:
                        dict_add_count[list_new[i].year] += 1
                    else:
                        dict_add_count[list_new[i].year] = 1  
                    #если есть столбец с цитированием
                    if "citation" in list_members:
                        #кол-во цитирований по годам
                        if list_new[i].year in dict_add_citated:
                            dict_add_citated[list_new[i].year] += int(list_new[i].citation)
                        else:
                            dict_add_citated[list_new[i].year] = int(list_new[i].citation)
                    #ставим линию в 1 и 2 колонке
                    ws.cell(row=i+temp_row, column=1).border = border
                    ws.cell(row=i+temp_row, column=2).border = border
                    count_article += 1 #кол-во записей
                    count += 1 #кол-во записей
                #по всем атрибутам класса
                for j in range(len(list_members)):
                    #если библиографический список, то делаем его
                    if list_members[j]=="full_bibliographic_description":
                        if not hasattr(list_new[i],'full_bibliographic_description'):
                            ws.cell(row=temp_row, column=j+3).value = full_description(list_new[i])
                        else:
                            ws.cell(row=temp_row, column=j+3).value = list_new[i].full_bibliographic_description
                    #иначе обычное поле заполняем
                    else:
                        ws.cell(row=temp_row, column=j+3).value = getattr(list_new[i], list_members[j])
                    #если статья другая, то ставим бордер
                    if repetition==False:
                        ws.cell(row=i+temp_row, column=j+3).border = border
                    #красим ячейку в зелёный
                    ws.cell(row=temp_row, column=j+3).fill = PatternFill(fill_type='solid', start_color='CCFFCC', end_color='CCFFCC')
            #нумеруем 1 и 2 колонку
            ws.cell(row=temp_row, column=1).value = count_article
            ws.cell(row=temp_row, column=2).value = count
            #меняем текущее название для дальнейшего бордера
            title = title_temp
            temp_row += 1


    count = 0
    #добавляем одинаковые элементы
    title = ""
    for i in range(len(list_ident)):
        #сотрудник
        employe_flag = False
        employe = list_ident[i].author
        for k in range(len(list_data)):
            if list_ident[i].author in list_data[k] or list_ident[i].author == list_data[k]:
                employe_flag = True
                break
        if employe_flag:
            #берём текущее название
            title_temp = getattr(list_ident[i], "title")
            #1 и 2 ячейки для нумерации
            ws.cell(row=temp_row, column=1).font = Font(size=16)
            ws.cell(row=temp_row, column=2).font = Font(size=16)
            #если текущее название не совпало с прошлым, то отделяем чертой ячейки
            if title_temp!= title:
                #заполяем словари
                #кол-во статей по годам
                if list_ident[i].year in dict_ident_count:
                    dict_ident_count[list_ident[i].year] += 1
                else:
                    dict_ident_count[list_ident[i].year] = 1  
            
                #если есть столбец с цитированием
                if "citation" in list_members:
                    #кол-во цитирований по годам
                    if list_ident[i].year in dict_ident_citated:
                        dict_ident_citated[list_ident[i].year] += int(list_ident[i].citation)
                    else:
                        dict_ident_citated[list_ident[i].year] = int(list_ident[i].citation)
                        
                #ставим линию в 1 и 2 колонке
                ws.cell(row=temp_row, column=1).border = border
                ws.cell(row=temp_row, column=2).border = border
                count_article += 1 #кол-во записей
                count += 1 #кол-во записей
                #Бежим по списку атрибутов класса
                for j in range(len(list_members)):
                    if list_members[j]=="full_bibliographic_description":
                        if not hasattr(list_ident[i],'full_bibliographic_description'):
                            ws.cell(row=temp_row, column=j+3).value = full_description(list_ident[i])
                        else:
                             ws.cell(row=temp_row, column=j+3).value = list_ident[i].full_bibliographic_description
                    #иначе обычное поле заполняем
                    else:
                        ws.cell(row=temp_row, column=j+3).value = getattr(list_ident[i], list_members[j])
                    #делаем бордер
                    ws.cell(row=temp_row, column=j+3).border = border
            #иначе если текущщее название свходится с прошлым, то бордер делать не надо
            else:
                repetition = True
                if hasattr(list_ident[i],'eid') and i>0 and list_ident[i].eid!=list_ident[i-1].eid:
                    repetition = False
                    #заполяем словари
                    #кол-во статей по годам
                    if list_ident[i].year in dict_add_count:
                        dict_add_count[list_ident[i].year] += 1
                    else:
                        dict_add_count[list_ident[i].year] = 1  
                    #если есть столбец с цитированием
                    if "citation" in list_members:
                        #кол-во цитирований по годам
                        if list_ident[i].year in dict_add_citated:
                            dict_add_citated[list_ident[i].year] += int(list_ident[i].citation)
                        else:
                            dict_add_citated[list_ident[i].year] = int(list_ident[i].citation)
                    #ставим линию в 1 и 2 колонке
                    ws.cell(row=i+temp_row, column=1).border = border
                    ws.cell(row=i+temp_row, column=2).border = border
                    count_article += 1 #кол-во записей
                    count += 1 #кол-во записей
                #по всем атрибутам класса
                for j in range(len(list_members)):
                    #если библиографический список, то делаем его
                    if list_members[j]=="full_bibliographic_description":
                        if not hasattr(list_ident[i],'full_bibliographic_description'):
                            ws.cell(row=temp_row, column=j+3).value = full_description(list_ident[i])
                        else:
                             ws.cell(row=temp_row, column=j+3).value = list_ident[i].full_bibliographic_description
                    #иначе обычное поле заполняем
                    else:
                        ws.cell(row=temp_row, column=j+3).value = getattr(list_ident[i], list_members[j])
                    #если статья другая, то ставим бордер
                    if repetition==False:
                        ws.cell(row=i+temp_row, column=j+3).border = border
            #нумеруем 1 и 2 колонку
            ws.cell(row=temp_row, column=1).value = count_article
            ws.cell(row=temp_row, column=2).value = count
            #меняем текущее название для дальнейшего бордера
            title = title_temp
            temp_row += 1


    count = 0
    #добавляем удалённые элементы
    title = ""
    for i in range(len(list_remove)):
        #сотрудник
        employe_flag = False
        employe = list_remove[i].author
        for k in range(len(list_data)):
            if list_remove[i].author in list_data[k] or list_remove[i].author == list_data[k]:
                employe_flag = True
                break
        if employe_flag:
            #берём текущее название
            title_temp = getattr(list_remove[i], "title")
            #1 и 2 ячейки для нумерации
            ws.cell(row=temp_row, column=1).font = Font(size=16)
            ws.cell(row=temp_row, column=2).font = Font(size=16)
            ws.cell(row=temp_row, column=1).fill = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
            ws.cell(row=temp_row, column=2).fill = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
            #если текущее название не совпало с прошлым, то отделяем чертой ячейки
            if title_temp!= title:
                #заполяем словари
                #кол-во статей по годам
                if list_remove[i].year in dict_remove_count:
                    dict_remove_count[list_remove[i].year] += 1
                else:
                    dict_remove_count[list_remove[i].year] = 1  
            
                #если есть столбец с цитированием
                if "citation" in list_members:
                    #кол-во цитирований по годам
                    if list_remove[i].year in dict_remove_citated:
                        dict_remove_citated[list_remove[i].year] += int(list_remove[i].citation)
                    else:
                        dict_remove_citated[list_remove[i].year] = int(list_remove[i].citation)
                        
                #ставим линию в 1 и 2 колонке
                ws.cell(row=temp_row, column=1).border = border
                ws.cell(row=temp_row, column=2).border = border
                count_article += 1 #кол-во записей
                count += 1 #кол-во записей
                #Бежим по списку атрибутов класса
                for j in range(len(list_members)):
                    if list_members[j]=="full_bibliographic_description":
                        if not hasattr(list_remove[i],'full_bibliographic_description'):
                            ws.cell(row=temp_row, column=j+3).value = full_description(list_remove[i])
                        else:
                            ws.cell(row=temp_row, column=j+3).value = list_remove[i].full_bibliographic_description
                    #иначе обычное поле заполняем
                    else:
                        ws.cell(row=temp_row, column=j+3).value = getattr(list_remove[i], list_members[j])
                    #делаем бордер и красим в красный цвет
                    ws.cell(row=temp_row, column=j+3).border = border
                    ws.cell(row=temp_row, column=j+3).fill = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
            #иначе если текущщее название свходится с прошлым, то бордер делать не надо
            else:
                repetition = True
                if hasattr(list_remove[i],'eid') and i>0 and list_remove[i].eid!=list_remove[i-1].eid:
                    repetition = False
                    #заполяем словари
                    #кол-во статей по годам
                    if list_remove[i].year in dict_add_count:
                        dict_add_count[list_remove[i].year] += 1
                    else:
                        dict_add_count[list_remove[i].year] = 1  
                    #если есть столбец с цитированием
                    if "citation" in list_members:
                        #кол-во цитирований по годам
                        if list_remove[i].year in dict_add_citated:
                            dict_add_citated[list_remove[i].year] += int(list_remove[i].citation)
                        else:
                            dict_add_citated[list_remove[i].year] = int(list_remove[i].citation)
                    #ставим линию в 1 и 2 колонке
                    ws.cell(row=i+temp_row, column=1).border = border
                    ws.cell(row=i+temp_row, column=2).border = border
                    count_article += 1 #кол-во записей
                    count += 1 #кол-во записей
                #по всем атрибутам класса
                for j in range(len(list_members)):
                    if list_members[j]=="full_bibliographic_description":
                        if not hasattr(list_remove[i],'full_bibliographic_description'):
                           ws.cell(row=temp_row, column=j+3).value = full_description(list_remove[i])
                        else:
                            ws.cell(row=temp_row, column=j+3).value = list_remove[i].full_bibliographic_description
                    #иначе обычное поле заполняем
                    else:
                        ws.cell(row=temp_row, column=j+3).value = getattr(list_remove[i], list_members[j])
                    #если статья другая, то ставим бордер
                    if repetition==False:
                        ws.cell(row=i+temp_row, column=j+3).border = border
                    #красим ячейку в красный
                    ws.cell(row=temp_row, column=j+3).fill = PatternFill(fill_type='solid', start_color='FFCCCC', end_color='FFCCCC')
            #нумеруем 1 и 2 колонку
            ws.cell(row=temp_row, column=1).value = count_article
            ws.cell(row=temp_row, column=2).value = count
            #меняем текущее название для дальнейшего бордера
            title = title_temp
            temp_row += 1

    #изменяем высоту ячейки
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 40

    #перенос строк
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
         for cell in row_cells:
            cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            cell.font = Font(size=16)
    
    #статистика в новый лист книги
    ws = wb.create_sheet('Статистика')
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30
    
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 30
    ws.column_dimensions["K"].width = 30
    ws.column_dimensions["L"].width = 30
    ws.column_dimensions["M"].width = 30
    ws.column_dimensions["N"].width = 30
    
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Кол-во статей по годам'
    
    ws.merge_cells('I1:N1')
    ws['I1'] = 'Кол-во цитирований по годам'
    
    ws.merge_cells('A2:B2')
    ws['A2'] = 'Добавленные статьи'
    
    ws.merge_cells('C2:D2')
    ws['C2'] = 'Одинаковые статьи'

    ws.merge_cells('E2:F2')
    ws['E2'] = 'Удалённые статьи'
    
    ws.merge_cells('I2:J2')
    ws['I2'] = 'Добавленные статьи'
    
    ws.merge_cells('K2:L2')
    ws['K2'] = 'Одинаковые статьи'

    ws.merge_cells('M2:N2')
    ws['M2'] = 'Удалённые статьи'
    
    ws['A3'] = 'Год'
    ws['C3'] = 'Год'
    ws['E3'] = 'Год'
    ws['I3'] = 'Год'
    ws['K3'] = 'Год'
    ws['M3'] = 'Год'
    
    ws['B3'] = 'Кол-во'
    ws['D3'] = 'Кол-во'
    ws['F3'] = 'Кол-во'
    ws['J3'] = 'Кол-во'
    ws['L3'] = 'Кол-во'
    ws['N3'] = 'Кол-во'
    
    i = 4
    for key in dict_add_count:
        ws.cell(row=i, column=1).value = key
        ws.cell(row=i, column=2).value = dict_add_count[key]
        i += 1
    
    i = 4
    for key in dict_ident_count:
        ws.cell(row=i, column=3).value = key
        ws.cell(row=i, column=4).value = dict_ident_count[key]
        i += 1

    i = 4
    for key in dict_remove_count:
        ws.cell(row=i, column=5).value = key
        ws.cell(row=i, column=6).value = dict_remove_count[key]
        i += 1
        
    i = 4
    for key in dict_add_citated:
        ws.cell(row=i, column=9).value = key
        ws.cell(row=i, column=10).value = dict_add_citated[key]
        i += 1
        
    i = 4
    for key in dict_ident_citated:
        ws.cell(row=i, column=11).value = key
        ws.cell(row=i, column=12).value = dict_ident_citated[key]
        i += 1
        
    i = 4
    for key in dict_remove_citated:
        ws.cell(row=i, column=13).value = key
        ws.cell(row=i, column=14).value = dict_remove_citated[key]
        i += 1
        

    #перенос строк и центирование текста
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
         for cell in row_cells:
            cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            cell.font = Font(size=16)
    
    wb.save(path[0:-5]+"_vyatsu.xlsx")