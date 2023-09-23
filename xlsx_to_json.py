from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from xls2xlsx import XLS2XLSX
from Translate import Translite
import json

filename = askopenfilename(filetypes = [("xlsx", "*.xlsx"), ('All files', '*')])
print(f"Расположение файла: {filename}\n")


try:
    if filename[-1]=="s":
        x2x = XLS2XLSX(filename)
        filename = filename + "x"
        x2x.to_xlsx(filename)
        
    #открываем файл
    wb = load_workbook(filename)
    ws = wb.active
    print("Файл успешно загружен\n")
    
    #формируем список сотрудников
    list_employee = []
    for row_index in range(14, ws.max_row + 1):
        font = ws[f"A{row_index}"].font
        if font.b==False:
            list_employee.append(ws[f"A{row_index}"].value)
       
    #преобразуем ФИО к новому виду
    new_list_employee = []
    for i in range(len(list_employee)):
        temp_list_employee = list_employee[i].split()
        string = f"{Translite(temp_list_employee[0])} {Translite(temp_list_employee[1])[0]}.{Translite(temp_list_employee[2])[0]}."
        new_list_employee.append(string)
        
    #убираем повторки
    list_employee = list(set(new_list_employee))
    list_employee.sort()
    
    #выгрузка в json файл
    json_string = json.dumps(list_employee, ensure_ascii=False, indent=4)
    json_dict = json.loads(json_string) 
    with open("employee.json", 'w', encoding="utf-8") as file:
        json.dump(json_dict, file, ensure_ascii=False, indent=4)
    print(json_string)

except:
    print("Возможно, Вы пытались загрузить файл с другим расширением, отличным от .xls или .xlsx")
    print("Программа закончилась аварийно")
    

print("\nЭто окно можно закрыть")
input()
