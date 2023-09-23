from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import json


def Upload(path, list_new = [], list_ident = [], list_remove = []):
    # создание нового файла
    wb = Workbook()
    ws = wb.active
    
    #границы ячейки
    border=Border(top=Side(border_style='thick',color="FFFF00"))

    #название листа 
    ws.title = "Сравнение данных"
        
    #получаем заголовки столбиков в список
    list_members = []
    if len(list_new)>0:
        list_members = list_new[0].__dir__()
        if len(list_ident)>0:
            list_members = list(set(list_members).intersection(set(list_ident[0].__dir__())))
        if len(list_remove)>0:
            list_members = list(set(list_members).intersection(set(list_remove[0].__dir__()))) 
    else:
        if len(list_ident)>0:
            list_members = list_ident[0].__dir__()
            if len(list_remove)>0:
                list_members = list(set(list_members).intersection(set(list_remove[0].__dir__()))) 
        else:
            if len(list_remove)>0:
                list_members = list_remove[0].__dir__()
    
    #удаляем спец. поля класса
    new_list_members = []       
    for var in list_members:
        if var[0] != "_" and var[0].islower() and var[:5]!="clear":
            new_list_members.append(var)
            
    #формируем сортированный список
    new_list_members.remove("author")
    new_list_members.remove("title")
    new_list_members.remove("year")
    list_members = ["№ article","author","title","year", "full bibliographic title"]
    
    #если есть поля, то добавляем в нужной очерёдности
    if "doi" in new_list_members:
        list_members.append("doi")
        new_list_members.remove("doi")
        
    if "link" in new_list_members:
        list_members.append("link")
        new_list_members.remove("link")
        
    if "start_page" in new_list_members:
        list_members.append("start_page")
        new_list_members.remove("start_page")
        
    if "end_page" in new_list_members:
        list_members.append("end_page")
        new_list_members.remove("end_page")
        
    if "number_of_pages" in new_list_members:
        list_members.append("number_of_pages")
        new_list_members.remove("number_of_pages")
        
    if "issue" in new_list_members:
        list_members.append("issue")
        new_list_members.remove("issue")
    
    if "volume" in new_list_members:
        list_members.append("volume")
        new_list_members.remove("volume")
        
    if "article" in new_list_members:
        list_members.append("article")
        new_list_members.remove("article")
        
    if "citation" in new_list_members:
        list_members.append("citation")
        new_list_members.remove("citation")
        
    #добавляем остальные поля
    new_list_members.sort()
    for i in range(len(new_list_members)):
        list_members.append(new_list_members[i])
        
    if "source" in new_list_members:
        list_members.remove("source")
        list_members.append("source")
        new_list_members.remove("source")

    #меняем заголовки столбиков
    for index, val in enumerate(list_members):
        ws.cell(row=1, column=index+1).value = val
        ws.cell(row=1, column=index+1).font = Font(size=20)
    
    #меняем ширину ячеек
    for i in range(len(list_members)):
        ws.column_dimensions[ws.cell(row=1, column=i+1).column_letter].width = 60
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 90
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 200
    ws.column_dimensions["F"].width = 70    
    ws.column_dimensions["G"].width = 150
      
    temp_row = 2
    count_article = 0
    
    #словарь с сотрудниками
    file = open('employee.json', 'r', encoding="utf-8")
    list_data = json.load(file)
    
    #добавляем новые элементы
    title = ""
    for i in range(len(list_new)):
        title_temp = getattr(list_new[i], "title")
        ws.cell(row=i+temp_row, column=1).font = Font(size=16)
        ws.cell(row=i+temp_row, column=1).fill = PatternFill(fill_type='solid', start_color='7FFF00', end_color='7FFF00')
        if title_temp!= title:
            ws.cell(row=i+temp_row, column=1).border = border
            count_article += 1
            for j in range(1, len(list_members)):
                if j==1:
                    for k in range(len(list_data)):
                        if list_new[i].author in list_data[k] or list_new[i].author == list_data[k]:
                            ws.cell(row=i+temp_row, column=j+1).border = Border(top = Side(border_style='thick', color='8B0000'),right = Side(border_style='thick', color='8B0000'),bottom = Side(border_style='thick', color='8B0000'),left = Side(border_style='thick', color='8B0000'))
                            break
                if j==4:
                    s = f"{list_new[i].author}, {list_new[i].title} - {list_new[i].year}."
                    if "volume" in list_members and list_new[i].volume != None:
                        s += f" - Vol. {list_new[i].volume}."
                    if "issue" in list_members and list_new[i].issue != None:
                        s += f" - №{list_new[i].issue}."
                    if "start_page" in list_members and "end_page" in list_members and list_new[i].start_page != None and list_new[i].end_page != None:
                        s += f" - pp. {list_new[i].start_page}-{list_new[i].end_page}"
                    if "source_name" in list_members and list_new[i].source_name != None:
                        s += f", {list_new[i].source_name}"
                    ws.cell(row=i+temp_row, column=j+1).value = s
                else:
                    ws.cell(row=i+temp_row, column=j+1).value = getattr(list_new[i], list_members[j])
                ws.cell(row=i+temp_row, column=j+1).border = border
                ws.cell(row=i+temp_row, column=j+1).fill = PatternFill(fill_type='solid', start_color='7FFF00', end_color='7FFF00')
        else:
            for j in range(1, len(list_members)):
                if j==1:
                    for k in range(len(list_data)):
                        if list_new[i].author in list_data[k] or list_new[i].author == list_data[k]:
                            ws.cell(row=i+temp_row, column=j+1).border = Border(top = Side(border_style='thick', color='8B0000'),right = Side(border_style='thick', color='8B0000'),bottom = Side(border_style='thick', color='8B0000'),left = Side(border_style='thick', color='8B0000'))
                            break
                if j==4:
                    s = f"{list_new[i].author}, {list_new[i].title} - {list_new[i].year}."
                    if "volume" in list_members and list_new[i].volume != None:
                        s += f" - Vol. {list_new[i].volume}."
                    if "issue" in list_members and list_new[i].issue != None:
                        s += f" - №{list_new[i].issue}."
                    if "start_page" in list_members and "end_page" in list_members and list_new[i].start_page != None and list_new[i].end_page != None:
                        s += f" - pp. {list_new[i].start_page}-{list_new[i].end_page}"
                    if "source_name" in list_members and list_new[i].source_name != None:
                        s += f", {list_new[i].source_name}"
                    ws.cell(row=i+temp_row, column=j+1).value = s
                else:
                    ws.cell(row=i+temp_row, column=j+1).value = getattr(list_new[i], list_members[j])
                ws.cell(row=i+temp_row, column=j+1).fill = PatternFill(fill_type='solid', start_color='7FFF00', end_color='7FFF00')
        ws.cell(row=i+temp_row, column=1).value = count_article
        title = title_temp
    temp_row += len(list_new)


    #добавляем одинаковые элементы
    title = ""
    for i in range(len(list_ident)):
        title_temp = getattr(list_ident[i], "title")
        ws.cell(row=i+temp_row, column=1).font = Font(size=16)
        if title_temp!= title:
            ws.cell(row=i+temp_row, column=1).border = border
            count_article += 1              
            for j in range(1, len(list_members)):
                if j==1:
                    for k in range(len(list_data)):
                        if list_ident[i].author in list_data[k] or list_ident[i].author == list_data[k]:
                            ws.cell(row=i+temp_row, column=j+1).border = Border(top = Side(border_style='thick', color='8B0000'),right = Side(border_style='thick', color='8B0000'),bottom = Side(border_style='thick', color='8B0000'),left = Side(border_style='thick', color='8B0000'))
                            break
                if j==4:
                    s = f"{list_ident[i].author}, {list_ident[i].title} - {list_ident[i].year}."
                    if "volume" in list_members and list_ident[i].volume != None:
                        s += f" - Vol. {list_ident[i].volume}."
                    if "issue" in list_members and list_ident[i].issue != None:
                        s += f" - №{list_ident[i].issue}."
                    if "start_page" in list_members and "end_page" in list_members and list_ident[i].start_page != None and list_ident[i].end_page != None:
                        s += f" - pp. {list_ident[i].start_page}-{list_ident[i].end_page}"
                    if "source_name" in list_members and list_ident[i].source_name != None:
                        s += f", {list_ident[i].source_name}"
                    ws.cell(row=i+temp_row, column=j+1).value = s
                else:
                    ws.cell(row=i+temp_row, column=j+1).value = getattr(list_ident[i], list_members[j]) 
                ws.cell(row=i+temp_row, column=j+1).border  = border
        else:
            for j in range(1, len(list_members)):
                if j==1:
                    for k in range(len(list_data)):
                        if list_ident[i].author in list_data[k] or list_ident[i].author == list_data[k]:
                            ws.cell(row=i+temp_row, column=j+1).border = Border(top = Side(border_style='thick', color='8B0000'),right = Side(border_style='thick', color='8B0000'),bottom = Side(border_style='thick', color='8B0000'),left = Side(border_style='thick', color='8B0000'))
                            break
                if j==4:
                    s = f"{list_ident[i].author}, {list_ident[i].title} - {list_ident[i].year}."
                    if "volume" in list_members and list_ident[i].volume != None:
                        s += f" - Vol. {list_ident[i].volume}."
                    if "issue" in list_members and list_ident[i].issue != None:
                        s += f" - №{list_ident[i].issue}."
                    if "start_page" in list_members and "end_page" in list_members and list_ident[i].start_page != None and list_ident[i].end_page != None:
                        s += f" - pp. {list_ident[i].start_page}-{list_ident[i].end_page}"
                    if "source_name" in list_members and list_ident[i].source_name != None:
                        s += f", {list_ident[i].source_name}"
                    ws.cell(row=i+temp_row, column=j+1).value = s
                else:
                    ws.cell(row=i+temp_row, column=j+1).value = getattr(list_ident[i], list_members[j])  
        ws.cell(row=i+temp_row, column=1).value = count_article
        title = title_temp
    temp_row += len(list_ident)
        

    #добавляем удалённые элементы
    title = ""
    for i in range(len(list_remove)):
        title_temp = getattr(list_remove[i], "title")
        ws.cell(row=i+temp_row, column=1).font = Font(size=16)
        ws.cell(row=i+temp_row, column=1).fill = PatternFill(fill_type='solid', start_color='F08080', end_color='F08080')
        if title_temp!= title:
            ws.cell(row=i+temp_row, column=1).border = border
            count_article += 1
            for j in range(1, len(list_members)):
                if j==1:
                    for k in range(len(list_data)):
                         if list_remove[i].author in list_data[k] or list_remove[i].author == list_data[k]:
                            ws.cell(row=i+temp_row, column=j+1).border = Border(top = Side(border_style='thick', color='8B0000'),right = Side(border_style='thick', color='8B0000'),bottom = Side(border_style='thick', color='8B0000'),left = Side(border_style='thick', color='8B0000'))
                            break
                if j==4:
                    s = f"{list_remove[i].author}, {list_remove[i].title} - {list_remove[i].year}."
                    if "volume" in list_members and list_remove[i].volume != None:
                        s += f" - Vol. {list_remove[i].volume}."
                    if "issue" in list_members and list_remove[i].issue != None:
                        s += f" - №{list_remove[i].issue}."
                    if "start_page" in list_members and "end_page" in list_members and list_remove[i].start_page != None and list_remove[i].end_page != None:
                        s += f" - pp. {list_remove[i].start_page}-{list_remove[i].end_page}"
                    if "source_name" in list_members and list_remove[i].source_name != None:
                        s += f", {list_remove[i].source_name}"
                    ws.cell(row=i+temp_row, column=j+1).value = s
                else:
                    ws.cell(row=i+temp_row, column=j+1).value = getattr(list_remove[i], list_members[j])
                ws.cell(row=i+temp_row, column=j+1).fill = PatternFill(fill_type='solid', start_color='F08080', end_color='F08080')
                ws.cell(row=i+temp_row, column=j+1).border  = border
        else:
            for j in range(1, len(list_members)):
                if j==1:
                    for k in range(len(list_data)):
                        if list_remove[i].author in list_data[k] or list_remove[i].author == list_data[k]:
                            ws.cell(row=i+temp_row, column=j+1).border = Border(top = Side(border_style='thick', color='8B0000'),right = Side(border_style='thick', color='8B0000'),bottom = Side(border_style='thick', color='8B0000'),left = Side(border_style='thick', color='8B0000'))
                            break
                if j==4:
                    s = f"{list_remove[i].author}, {list_remove[i].title} - {list_remove[i].year}."
                    if "volume" in list_members and list_remove[i].volume != None:
                        s += f" - Vol. {list_remove[i].volume}."
                    if "issue" in list_members and list_remove[i].issue != None:
                        s += f" - №{list_remove[i].issue}."
                    if "start_page" in list_members and "end_page" in list_members and list_remove[i].start_page != None and list_remove[i].end_page != None:
                        s += f" - pp. {list_remove[i].start_page}-{list_remove[i].end_page}"
                    if "source_name" in list_members and list_remove[i].source_name != None:
                        s += f", {list_remove[i].source_name}"
                    ws.cell(row=i+temp_row, column=j+1).value = s
                else:
                    ws.cell(row=i+temp_row, column=j+1).value = getattr(list_remove[i], list_members[j])
                ws.cell(row=i+temp_row, column=j+1).fill = PatternFill(fill_type='solid', start_color='F08080', end_color='F08080')
        ws.cell(row=i+temp_row, column=1).value = count_article
        title = title_temp
    temp_row += len(list_remove)
             
    #изменяем высоту ячейки
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 40
        
    #перенос строк
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
         for cell in row_cells:
            cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            cell.font = Font(size=16)
            
    print(f"Для проверки загрузки: {len(list_new)+len(list_ident)+len(list_remove)} = {ws.max_row-1}")
    wb.save(path)
