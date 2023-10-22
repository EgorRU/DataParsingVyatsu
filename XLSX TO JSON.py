from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from xls2xlsx import XLS2XLSX
import json
import traceback


def Translite(s_input):
    s_output = ""
    if s_input!=None and len(s_input)>0:
        for i in range(len(s_input)):
            if s_input[i] in translate_dictionary:
                s_output += translate_dictionary[s_input[i]]
            else:
                s_output += s_input[i]
    return s_output


translate_dictionary = {
  "а": "a",
  "б": "b",
  "в": "v",
  "г": "g",
  "д": "d",
  "е": "e",
  "ё": "yo",
  "ж": "zh",
  "з": "z",
  "и": "i",
  "й": "j",
  "к": "k",
  "л": "l",
  "м": "m",
  "н": "n",
  "о": "o",
  "п": "p",
  "р": "r",
  "с": "s",
  "т": "t",
  "у": "u",
  "ф": "f",
  "х": "kh",
  "ц": "ts",
  "ч": "ch",
  "ш": "sh",
  "щ": "shch",
  "ь": "",
  "ы": "y",
  "ъ": "",
  "э": "e",
  "ю": "yu",
  "я": "ya",
  "А": "A",
  "Б": "B",
  "В": "V",
  "Г": "G",
  "Д": "D",
  "Е": "E",
  "Ё": "Yo",
  "Ж": "Zh",
  "З": "Z",
  "И": "I",
  "Й": "J",
  "К": "K",
  "Л": "L",
  "М": "M",
  "Н": "N",
  "О": "O",
  "П": "P",
  "Р": "R",
  "С": "S",
  "Т": "T",
  "У": "U",
  "Ф": "F",
  "Х": "Kh",
  "Ц": "Ts",
  "Ч": "Ch",
  "Ш": "Sh",
  "Щ": "Shch",
  "Ы": "Y",
  "Э": "E",
  "Ю": "Yu",
  "Я": "Ya"
}


filename = askopenfilename(filetypes = [("xlsx, xls", "*.xls?"), ('All files', '*')])
print(f"Расположение файла: {filename}\n")

try:
    if filename[-1]=="s":
        x2x = XLS2XLSX(filename)
        filename = filename + "x"
        x2x.to_xlsx(filename)
        
    #открываем файл
    wb = load_workbook(filename)
    ws = wb.active
    print("Файл успешно загружен\n")
    
    #формируем список сотрудников
    list_employee = []
    for row_index in range(14, ws.max_row + 1):
        font = ws[f"A{row_index}"].font
        if font.b==False:
            list_employee.append(ws[f"A{row_index}"].value)
       
    #преобразуем ФИО к новому виду
    new_list_employee = []
    for i in range(len(list_employee)):
        try:
            temp_list_employee = list_employee[i].split()
            string = f"{Translite(temp_list_employee[0])} {Translite(temp_list_employee[1])[0]}."
            if len(temp_list_employee)==3:
                string += f"{Translite(temp_list_employee[2])[0]}."
            new_list_employee.append(string)
        except Exception as e:
            print(str(e))
            print(traceback.format_exc())
            print(f"Сломанный автор: {list_employee[i]}")
        
    #убираем повторки
    list_employee = list(set(new_list_employee))
    list_employee.sort()
    
    #выгрузка в json файл
    json_string = json.dumps(list_employee, ensure_ascii=False, indent=4)
    json_dict = json.loads(json_string) 
    with open("employee.json", 'w', encoding="utf-8") as file:
        json.dump(json_dict, file, ensure_ascii=False, indent=4)
    print(json_string)
    
except Exception as e:
    print(str(e))
    print(traceback.format_exc())
    print("[!]Программа завершилась аварийно")
    
print("\n[!]Это окно можно закрыть")
input()
