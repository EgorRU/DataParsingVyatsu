from openpyxl import load_workbook
from App import IPublishing_Library
from App import Translite
from App import clear_author
from App import clear_IPublishing_title


def IPublishing(path):
    # список статей IPublishing
    all_IPublishing_list_library = []
    #открытие исходного файла с проверкой 
    try:
        wb = load_workbook(path)
    except:
        return None

    #парсим 1-й лист
    ws = wb["1. Статьи в журналах"]
    # ПАРСИНГ КАЖДОЙ СТРОКИ
    for row_index in range(5, ws.max_row + 1):
        try:
            str_author = ws[f"G{row_index}"].value #строка с авторами 
            list_author = str_author.split(",") #список авторов
            for i in range (len(list_author)):
                if list_author[i].strip() !="":
                    new_author = IPublishing_Library()
                    new_author.original_author = list_author[i].strip()
            
                    if ws[f"H{row_index}"].value != None:
                        new_author.title = ws[f"H{row_index}"].value
            
                    if ws[f"E{row_index}"].value != None:
                        new_author.year = ws[f"E{row_index}"].value
            
                    if ws[f"I{row_index}"].value != None:
                        new_author.title_article = ws[f"I{row_index}"].value
            
                    if ws[f"AF{row_index}"].value != None:
                        new_author.link = ws[f"AF{row_index}"].value
            
                    if ws[f"AG{row_index}"].value != None:
                        new_author.doi = ws[f"AG{row_index}"].value
            
                    if ws[f"B{row_index}"].value != None:
                        new_author.institute = ws[f"B{row_index}"].value
            
                    if ws[f"C{row_index}"].value != None:
                        new_author.faculty = ws[f"C{row_index}"].value
            
                    if ws[f"D{row_index}"].value != None:
                        new_author.cathedra = ws[f"D{row_index}"].value
                
                    if ws[f"J{row_index}"].value != None:
                        new_author.full_bibliographic_description = ws[f"J{row_index}"].value
                
                    if ws[f"V{row_index}"].value != None:
                        new_author.cod_OECD = ws[f"V{row_index}"].value
                
                    if ws[f"W{row_index}"].value != None:
                        new_author.group_of_scientific_specialties = ws[f"W{row_index}"].value
                
                    if ws[f"X{row_index}"].value != None:
                        new_author.GRNTI_code = ws[f"X{row_index}"].value
                
                    if ws[f"Y{row_index}"].value != None:
                        new_author.quartile_wos = ws[f"Y{row_index}"].value
                
                    if ws[f"Z{row_index}"].value != None:
                        new_author.quartile_scopus = ws[f"Z{row_index}"].value
                
                    if ws[f"AA{row_index}"].value != None:
                        new_author.quartile_scopus_sjr = ws[f"AA{row_index}"].value
                
                    if ws[f"AB{row_index}"].value != None:
                        new_author.impact_factor_wos = ws[f"AB{row_index}"].value
                
                    if ws[f"AC{row_index}"].value != None:
                        new_author.impact_factor_scopus = ws[f"AC{row_index}"].value
                    
                    if ws[f"AD{row_index}"].value != None:
                        new_author.impact_factor_elibrary_5_year = ws[f"AD{row_index}"].value
                
                    if ws[f"AE{row_index}"].value != None:
                        new_author.impact_factor_elibrary_2_year = ws[f"AE{row_index}"].value
                    new_author.source = "Статьи в журналах"
                    all_IPublishing_list_library.append(new_author)
        except:
            pass


    #парсим 2-й лист
    ws = wb["2. Публ. в научн. сборниках"]
    # ПАРСИНГ КАЖДОЙ СТРОКИ
    for row_index in range(5, ws.max_row + 1):
        try:
            str_author = ws[f"F{row_index}"].value #строка с авторами 
            list_author = str_author.split(",") #список авторов
            for i in range (len(list_author)):
                if list_author[i].strip() !="":
                    new_author = IPublishing_Library()
                    #формируем фио
                    new_author.original_author = list_author[i].strip()
                    index = new_author.original_author.find("(")
                    if index != -1:
                        new_author.original_author = new_author.original_author[:index].strip()
                
                    if ws[f"G{row_index}"].value != None:
                        new_author.title = clear_IPublishing_title(ws[f"G{row_index}"].value)
                        new_author.full_bibliographic_description = ws[f"G{row_index}"].value                
                    
                    if ws[f"E{row_index}"].value != None:
                        new_author.year = ws[f"E{row_index}"].value
            
                    if ws[f"H{row_index}"].value != None:
                        new_author.title_article = ws[f"H{row_index}"].value
            
                    if ws[f"R{row_index}"].value != None:
                        new_author.link = ws[f"R{row_index}"].value
            
                    if ws[f"S{row_index}"].value != None:
                        new_author.doi = ws[f"S{row_index}"].value
            
                    if ws[f"B{row_index}"].value != None:
                        new_author.institute = ws[f"B{row_index}"].value
            
                    if ws[f"C{row_index}"].value != None:
                        new_author.faculty = ws[f"C{row_index}"].value
            
                    if ws[f"D{row_index}"].value != None:
                        new_author.cathedra = ws[f"D{row_index}"].value
                
                    if ws[f"O{row_index}"].value != None:
                        new_author.cod_OECD = ws[f"O{row_index}"].value
                
                    if ws[f"P{row_index}"].value != None:
                        new_author.group_of_scientific_specialties = ws[f"P{row_index}"].value
                
                    if ws[f"Q{row_index}"].value != None:
                        new_author.GRNTI_code = ws[f"Q{row_index}"].value
                
                    new_author.source = "Публ. в научн. сборниках"
                    all_IPublishing_list_library.append(new_author)
        except:
            pass
    

    #парсим 3-й лист
    ws = wb["3. Монографии"]
    # ПАРСИНГ КАЖДОЙ СТРОКИ
    for row_index in range(5, ws.max_row + 1):
        try:    
            str_author = ws[f"G{row_index}"].value #строка с авторами 
            list_author = str_author.split(",") #список авторов
            for i in range (len(list_author)):
                if list_author[i].strip() !="":
                    new_author = IPublishing_Library()
                    new_author.original_author = list_author[i].strip()
            
                    if ws[f"H{row_index}"].value != None:
                        new_author.title = ws[f"H{row_index}"].value
            
                    if ws[f"E{row_index}"].value != None:
                        new_author.year = ws[f"E{row_index}"].value
            
                    if ws[f"T{row_index}"].value != None:
                        new_author.link = ws[f"T{row_index}"].value
            
                    if ws[f"I{row_index}"].value != None:
                        new_author.full_bibliographic_description = ws[f"I{row_index}"].value
            
                    if ws[f"B{row_index}"].value != None:
                        new_author.institute = ws[f"B{row_index}"].value
            
                    if ws[f"C{row_index}"].value != None:
                        new_author.faculty = ws[f"C{row_index}"].value
            
                    if ws[f"D{row_index}"].value != None:
                        new_author.cathedra = ws[f"D{row_index}"].value
                
                    if ws[f"O{row_index}"].value != None:
                        new_author.cod_OECD = ws[f"O{row_index}"].value
                
                    if ws[f"P{row_index}"].value != None:
                        new_author.group_of_scientific_specialties = ws[f"P{row_index}"].value
                
                    if ws[f"Q{row_index}"].value != None:
                        new_author.GRNTI_code = ws[f"Q{row_index}"].value
                
                    new_author.source = "Монографии"
                    all_IPublishing_list_library.append(new_author)
        except:
            pass

    for i in range(len(all_IPublishing_list_library)):
        try:
            all_IPublishing_list_library[i].author = clear_author(Translite(clear_author(all_IPublishing_list_library[i].original_author)))
        except:
            all_IPublishing_list_library[i].author = all_IPublishing_list_library[i].original_author
        all_IPublishing_list_library[i].author = all_IPublishing_list_library[i].author.replace("Bajkova", "Baykova")

    for i in range(len(all_IPublishing_list_library)):
        all_IPublishing_list_library[i].clear_title = "".join(e for e in all_IPublishing_list_library[i].title.lower() if e.isalpha())
        
    for i in range(len(all_IPublishing_list_library)):
        all_IPublishing_list_library[i].clear_author = "".join(e for e in all_IPublishing_list_library[i].author if e.isupper())
        
    return all_IPublishing_list_library
