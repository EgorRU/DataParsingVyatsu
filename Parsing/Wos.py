from openpyxl import load_workbook
from xls2xlsx import XLS2XLSX
import logging
from App import WOS_Library
from App import clear_author


LOG_FILENAME = 'log.out'
logging.basicConfig(filename=LOG_FILENAME, level=logging.DEBUG,  encoding='utf-8', format='%(asctime)s:%(levelname)s:%(message)s')


def Wos(path):
    # список статей Wos
    all_wos_list_library = []
    # открытие исходного файла c проверкой
    try:
        if path[-1]=="s":
            x2x = XLS2XLSX(path)
            path = path + "x"
            x2x.to_xlsx(path)
        #открываем файл
        wb = load_workbook(path)
        ws = wb.active
    except Exception as e: 
        logging.exception(str(e))
        return None

    #ПАРСИНГ КАЖДОЙ СТРОКИ
    for row_index in range(2, ws.max_row + 1):
        try:    
            str_author = ws[f"F{row_index}"].value #строка с авторами 
            str_author = str_author.replace(",","") #убираем запятые
            list_author = str_author.split(";") #список авторов
            for i in range (len(list_author)):
                new_author = WOS_Library()
                new_author.original_author = list_author[i].strip()
                try:
                    new_author.author = clear_author(list_author[i].strip())
                except Exception as e: 
                    logging.exception(str(e))
                    logging.info(f"Сломанный автор в восе: {list_author[i]}")
                    new_author.author = list_author[i].strip()
            
                if ws[f"I{row_index}"].value != None:
                    new_author.title = ws[f"I{row_index}"].value.lower()
                    new_author.title = new_author.title[0].upper() + new_author.title[1:]
                    
                if ws[f"J{row_index}"].value != None:
                    new_author.title_article = ws[f"J{row_index}"].value
                    
                if ws[f"BD{row_index}"].value != None:
                    new_author.number_article = ws[f"BD{row_index}"].value
                
                if ws[f"O{row_index}"].value != None:
                    new_author.conference_title = ws[f"O{row_index}"].value
                
                if ws[f"P{row_index}"].value != None:
                    new_author.conference_date = ws[f"P{row_index}"].value
                
                if ws[f"Q{row_index}"].value != None:
                    new_author.conference_location = ws[f"Q{row_index}"].value
                   
                if ws[f"AU{row_index}"].value != None:
                    new_author.year = ws[f"AU{row_index}"].value
                
                if ws[f"AV{row_index}"].value != None:
                    new_author.volume = ws[f"AV{row_index}"].value
                
                if ws[f"AW{row_index}"].value != None:
                    new_author.issue = ws[f"AW{row_index}"].value            
                
                if ws[f"BB{row_index}"].value != None:
                    new_author.start_page = ws[f"BB{row_index}"].value
                
                if ws[f"BC{row_index}"].value != None:
                    new_author.end_page = ws[f"BC{row_index}"].value
            
                if ws[f"BE{row_index}"].value != None:
                    new_author.doi = ws[f"BE{row_index}"].value
            
                if ws[f"BF{row_index}"].value != None:
                    new_author.link = ws[f"BF{row_index}"].value[10:].strip()
                
                if ws[f"AA{row_index}"].value != None:
                    new_author.researcher_ids = ws[f"AA{row_index}"].value
                
                if ws[f"AB{row_index}"].value != None:
                    new_author.ORCIDs = ws[f"AB{row_index}"].value
                
                if ws[f"AO{row_index}"].value != None:
                    new_author.issn = ws[f"AO{row_index}"].value
                
                if ws[f"AP{row_index}"].value != None:
                    new_author.eissn = ws[f"AP{row_index}"].value
                
                if ws[f"BS{row_index}"].value != None:
                    new_author.unique_wos_id = ws[f"BS{row_index}"].value[4:]
                    
                if ws[f"N{row_index}"].value != None:
                    new_author.document_type = ws[f"N{row_index}"].value
                    
                if ws[f"AF{row_index}"].value != None:
                    new_author.citations = ws[f"AF{row_index}"].value
                
                all_wos_list_library.append(new_author)
        except Exception as e: 
            logging.exception(str(e))
            logging.info(f"Сломанная строка в восе: {row_index}")

    for i in range(len(all_wos_list_library)):
        all_wos_list_library[i].clear_title = "".join(e for e in all_wos_list_library[i].title.lower() if e.isalpha())
        all_wos_list_library[i].clear_author = "".join(e for e in all_wos_list_library[i].author if e.isupper())
        all_wos_list_library[i].title = all_wos_list_library[i].title.replace('ё', 'e')
        
    return all_wos_list_library
