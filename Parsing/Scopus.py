from openpyxl import Workbook, load_workbook
from App import Scopus_Library
from App import clear_author


def Scopus(path):
    # список статей SCOPUS
    all_scopus_list_library = []

    # открытие исходного файла c проверкой
    try:
        wb = load_workbook(path)
        ws = wb.active
    except:
        return None

    # создание нового файла
    wb_new = Workbook()
    ws_new = wb_new.active

    # конкатенация строк
    count = 1
    for row in ws.iter_rows(values_only=True):
        s = ""
        for data in row:
            if data is not None:
                s += str(data)
                s += " "
        # запись в ячейку первой строки
        ws_new.cell(row=count, column=1).value = s
        count += 1

    # проходимся по новому файлу
    wb = wb_new
    ws = wb.active

    # кол-во ФИО авторов
    count_all_author = 0

    # ПАРСИНГ КАЖДОЙ СТРОКИ
    for row in ws.iter_rows(min_row=2, values_only=True):
        count_temp = count_all_author
        s = row[0]  # текущая строка
        i = 0  # индекс  - бежим по строке

        # кол-во авторов статьи в одной строке
        count_author_row = 0
        try:
            #ПАРСИНГ ФИО
            while s[i] != '"':
                # обнуляем строку имени писателя
                author = ""
                # пока не дошли до запятой - до конца ФИО писателя
                while s[i] != ",":
                    author += s[i]
                    i += 1
                # создание экземпляра класса - статья scopus
                new_author = Scopus_Library()
                count_author_row += 1
                new_author.original_author = author.strip()
                # заполянем ФИО автора
                try:
                    new_author.author = clear_author(author.strip())
                except:
                    new_author.author = author.strip()
                
                # добавляем экземпляр класса
                all_scopus_list_library.append(new_author)
                i += 1
            i += 1
            count_all_author += count_author_row

            #ПАРСИНГ id_author
            count_temp_id = count_temp
            # пока не буква, то есть начало названия
            while s[i].isalpha() == False:
                while s[i] == " ":
                    i += 1
                id_author = ""
                # пока цифра - идентификатор автора
                while s[i].isdigit():
                    id_author += s[i]
                    i += 1
                if id_author != "":
                    all_scopus_list_library[count_temp_id].id_author = id_author.strip()
                    count_temp_id += 1
                else:
                    break

            #ПАРСИНГ НАЗВАНИЯ
            while s[i].isalpha() == False:
                i += 1
            title = ""
            while s[i] != '"':
                title += s[i]
                i += 1
                if s[i] == '"' and s[i+1]=='"':
                    i+=2
            title = title.strip()
            #чистим название
            #если в конце скобки, то убираем их
            if title[-1]==")" or title[-1]=="]":
                #инвертируем строку
                s_temp = title[::-1]
                i_temp = 0
                if s_temp[i_temp]==")":
                    i_temp += 1
                    while s_temp[i_temp]!="(":
                        i_temp += 1
                elif s_temp[i_temp]=="]":
                    i_temp += 1
                    while s_temp[i_temp]!="[":
                        i_temp += 1
                i_temp += 1
                s_temp = s_temp[i_temp:]
                title = s_temp[::-1]
            for j in range(count_temp, count_temp + count_author_row):
                all_scopus_list_library[j].title = title

            #ПАРСИНГ ГОДА
            while s[i].isdigit() == False:
                i += 1
            year = ""
            while s[i] != ",":
                year += s[i]
                i += 1
            for j in range(count_temp, count_temp + count_author_row):
                all_scopus_list_library[j].year = year.strip()[:4]

            #ПАРСИНГ НАЗВАНИЯ ИСТОЧНИКА
            if s[i + 1] != ",":
                while s[i].isalpha() == False:
                    i += 1
                source_name = ""
                while s[i] != '"':
                    source_name += s[i]
                    i += 1
                    if s[i] == '"' and s[i + 1] == '"':
                        source_name += '""'
                        i += 2
                for j in range(count_temp, count_temp + count_author_row):
                    all_scopus_list_library[j].source_name = source_name.strip()
            i += 1

            #ПАРСИНГ НОМЕРА ТОМА
            #если следующий символ НЕ запятая, то значение ЕСТЬ(или кавычка стоит)
            if s[i + 1] != ",":
                i += 1
                is_apostrophes = False
                if s[i] == '"':
                    i += 1
                    is_apostrophes = True
                volume = ""
                if is_apostrophes:
                    while s[i] != '"':
                        volume += s[i]
                        i += 1
                    i += 1
                else:
                    while s[i].isdigit():
                        volume += s[i]
                        i += 1
                if volume != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].volume = volume.strip()
            else:
                i += 1

            #ПАРСИНГ НОМЕРА ВЫПУСКа
            if s[i + 1] != ",":
                i += 1
                is_apostrophes = False
                if s[i] == '"':
                    i += 1
                    is_apostrophes = True
                issue = ""
                if is_apostrophes:
                    while s[i] != '"':
                        issue += s[i]
                        i += 1
                    i += 1
                else:
                    while s[i].isdigit():
                        issue += s[i]
                        i += 1
                if issue != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].issue = issue.strip()
            else:
                i += 1

            #ПАРСИНГ НОМЕРА СТАТЬИ
            if s[i + 1] != ",":
                i += 1
                is_apostrophes = False
                if s[i] == '"':
                    i += 1
                    is_apostrophes = True
                number_article = ""
                if is_apostrophes:
                    while s[i] != '"':
                        number_article += s[i]
                        i += 1
                    i += 1
                else:
                    while s[i].isdigit():
                        number_article += s[i]
                        i += 1
                if number_article != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].number_article = number_article.strip()
            else:
                i += 1

            #ПАРСИНГ НОМЕРА СТРАНИЦЫ НАЧАЛА
            if s[i + 1] != ",":
                i += 1
                is_apostrophes = False
                if s[i] == '"':
                    i += 1
                    is_apostrophes = True
                start_page = ""
                if is_apostrophes:
                    while s[i] != '"':
                        start_page += s[i]
                        i += 1
                    i += 1
                else:
                    while s[i].isdigit():
                        start_page += s[i]
                        i += 1
                if start_page != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].start_page = start_page.strip()
            else:
                i += 1

            #ПАРСИНГ НОМЕРА СТРАНИЦЫ ОКОНЧАНИЯ
            if s[i + 1] != ",":
                i += 1
                is_apostrophes = False
                if s[i] == '"':
                    i += 1
                    is_apostrophes = True
                end_page = ""
                if is_apostrophes:
                    while s[i] != '"':
                        end_page += s[i]
                        i += 1
                    i += 1
                else:
                    while s[i].isdigit():
                        end_page += s[i]
                        i += 1
                if end_page != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].end_page = end_page.strip()
            else:
                i += 1

            #ПАРСИНГ КОЛИЧЕСТВО СТРАНИЦ
            if s[i + 1] != ",":
                i += 1
                is_apostrophes = False
                if s[i] == '"':
                    i += 1
                    is_apostrophes = True
                number_of_pages = ""
                if is_apostrophes:
                    while s[i] != '"':
                        number_of_pages += s[i]
                        i += 1
                    i += 1
                else:
                    while s[i].isdigit():
                        number_of_pages += s[i]
                        i += 1
                if number_of_pages != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].number_of_pages = number_of_pages.strip()
            else:
                i += 1

            #ПАРСИНГ Цитирования
            if s[i + 1] != ",":
                i += 1
                is_apostrophes = False
                if s[i] == '"':
                    i += 1
                    is_apostrophes = True
                citation = ""
                if is_apostrophes:
                    while s[i] != '"':
                        citation += s[i]
                        i += 1
                    i += 1
                else:
                    while s[i].isdigit():
                        citation += s[i]
                        i += 1
                if citation != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].citation = citation.strip()
            else:
                i += 1

            #ПАРСИНГ DOI
            if s[i + 1] != ",":
                i += 2
                doi = ""
                while s[i] != '"':
                    doi += s[i]
                    i += 1
                if doi != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].doi = doi.strip()
            i += 1

            #ПАРСИНГ Ссылка
            if s[i + 1] != ",":
                i += 1
                while s[i] == '"':
                    i += 1
                link = ""
                while s[i] != '"':
                    link += s[i]
                    i += 1
                if link != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].link = link.strip()
            i += 1
            
            #парсинг EID
            new_s = s[::-1]
            eid = ""
            index = 0
            while(new_s[index]!=","):
                eid += new_s[index]
                index +=1
            if eid != "":
                for j in range(count_temp, count_temp + count_author_row):
                    all_scopus_list_library[j].eid = eid[::-1].strip()
            index += 1
            
            #парсинг источника
            source = ""
            while(new_s[index]!=","):
                source += new_s[index]
                index +=1
            if source != "":
                for j in range(count_temp, count_temp + count_author_row):
                    all_scopus_list_library[j].source = source[::-1].strip()
            index += 1
            
            #парсинг доступа книги
            access = ""
            #если попались кавычки - то есть поле
            if new_s[index]=="\"":
                index += 1
                #если поле не пустое
                if new_s[index]!="\"":
                    while(new_s[index]!="\""):
                        access += new_s[index]
                        index +=1
                if access != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].access = access[::-1].strip()
                    index += 1
                else:
                    index += 1
                
            #парсинг publication_stage
            publication_stage = ""
            index += 1
            #если попались кавычки - то есть поле 
            if new_s[index]=="\"":
                index += 1
                #если поле не пустое
                if new_s[index]!="\"":
                    while(new_s[index]!="\""):
                        publication_stage += new_s[index]
                        index +=1
                if publication_stage != "":
                    for j in range(count_temp, count_temp + count_author_row):
                        all_scopus_list_library[j].publication_stage = publication_stage[::-1].strip()
                    index += 1
                else:
                    index += 1
            index += 1
                
            #парсинг type_document
            type_document = ""
            while(new_s[index]!=","):
                type_document += new_s[index]
                index +=1
            if type_document != "":
                for j in range(count_temp, count_temp + count_author_row):
                    all_scopus_list_library[j].type_document = type_document[::-1].strip()
            index += 2
            
            #парсинг языка
            lang = ""
            while(new_s[index]!="\""):
                lang += new_s[index]
                index +=1
            if lang != "":
                for j in range(count_temp, count_temp + count_author_row):
                    if len(lang)<20:
                        all_scopus_list_library[j].lang = lang[::-1].strip()
        except:
            pass

    for i in range(len(all_scopus_list_library)):
        all_scopus_list_library[i].clear_title = "".join(e for e in all_scopus_list_library[i].title.lower() if e.isalpha())
        all_scopus_list_library[i].clear_author = "".join(e for e in all_scopus_list_library[i].author if e.isupper())
        all_scopus_list_library[i].title = all_scopus_list_library[i].title.replace('ё', 'e')
    
    return all_scopus_list_library
